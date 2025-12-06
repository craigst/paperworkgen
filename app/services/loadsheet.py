"""
Loadsheet generation service.

Handles Excel template population and PDF conversion for loadsheets.
"""

import logging
from datetime import datetime
from typing import List, Optional

from openpyxl import load_workbook

from ..config import get_settings
from ..schemas import CarModel, LoadsheetRequest, GenerateResponse
from .helpers import (
    add_signature,
    convert_excel_to_pdf,
    ensure_output_folder,
    format_date_for_cell,
    format_folder_date,
    get_week_end_from_date,
    sanitize_filename,
    select_signature,
)

logger = logging.getLogger("paperworkgen.services.loadsheet")
settings = get_settings()

CAR_CELL_MAP = [
    {"make_model": "B11", "reg": "B13", "offloaded": "E10", "docs": "G10", "spare_keys": "I10", "notes": "C11"},
    {"make_model": "B15", "reg": "B17", "offloaded": "E14", "docs": "G14", "spare_keys": "I14", "notes": "C15"},
    {"make_model": "B19", "reg": "B21", "offloaded": "E18", "docs": "G18", "spare_keys": "I18", "notes": "C19"},
    {"make_model": "B23", "reg": "B25", "offloaded": "E22", "docs": "G22", "spare_keys": "I22", "notes": "C23"},
    {"make_model": "B27", "reg": "B29", "offloaded": "E26", "docs": "G26", "spare_keys": "I26", "notes": "C27"},
    {"make_model": "B31", "reg": "B33", "offloaded": "E30", "docs": "G30", "spare_keys": "I30", "notes": "C31"},
    {"make_model": "B35", "reg": "B37", "offloaded": "E34", "docs": "G34", "spare_keys": "I34", "notes": "C35"},
    {"make_model": "B39", "reg": "B41", "offloaded": "E38", "docs": "G38", "spare_keys": "I38", "notes": "C39"},
]


def _generate_load_summary(cars: List[CarModel]) -> str:
    """Generate textual summary that matches the legacy script."""
    not_offloaded = sum(1 for car in cars if car.offloaded.upper() != "Y")
    with_docs = sum(1 for car in cars if car.docs.upper() == "Y")
    with_spare = sum(1 for car in cars if car.spare_keys.upper() == "Y")

    summary_parts = [f"{not_offloaded} CARS LOADED"]
    if with_docs:
        summary_parts.append(f"{with_docs} CARS HAVE DOCUMENTS AND HAVE BEEN PLACED ON THE PASSENGER SEAT")
    else:
        summary_parts.append("0 CARS HAVE DOCUMENTS")
    summary_parts.append(f"{with_spare} CARS HAVE SPARE KEYS")
    return ", ".join(summary_parts)


def generate_loadsheet(request: LoadsheetRequest) -> GenerateResponse:
    """Generate loadsheet Excel and PDF files."""

    if not settings.loadsheet_template.exists():
        raise FileNotFoundError(f"Loadsheet template not found: {settings.loadsheet_template}")

    if len(request.cars) > len(CAR_CELL_MAP):
        raise ValueError(f"Loadsheet supports up to {len(CAR_CELL_MAP)} cars per template")

    load_date = datetime.strptime(request.load_date, "%Y-%m-%d")
    week_end = get_week_end_from_date(load_date)
    folder = ensure_output_folder(week_end)

    safe_collection = sanitize_filename(request.collection_point)
    filename = f"{request.load_number}_{safe_collection}"
    excel_path = folder / f"{filename}.xlsx"

    wb = load_workbook(settings.loadsheet_template)
    ws = wb.active

    ws["C6"] = format_date_for_cell(load_date)
    ws["G6"] = request.load_number
    ws["C7"] = request.fleet_reg.upper()
    ws["B9"] = request.collection_point.upper()
    ws["F9"] = request.delivery_point.upper()
    ws["C46"] = format_date_for_cell(load_date)
    ws["H46"] = format_date_for_cell(load_date)

    for idx, mapping in enumerate(CAR_CELL_MAP):
        if idx < len(request.cars):
            car = request.cars[idx]
            ws[mapping["make_model"]] = car.make_model.upper()
            ws[mapping["reg"]] = car.reg.upper()
            ws[mapping["offloaded"]] = car.offloaded.upper()
            ws[mapping["docs"]] = car.docs.upper()
            ws[mapping["spare_keys"]] = car.spare_keys.upper()
            ws[mapping["notes"]] = car.car_notes.upper()
        else:
            for cell in mapping.values():
                ws[cell] = ""

    summary = _generate_load_summary(request.cars)
    notes = [summary]
    if request.load_notes.strip():
        notes.append(request.load_notes.strip())
    for car in request.cars:
        if car.car_notes.strip():
            notes.append(car.car_notes.upper())

    ws["C39"] = "\n".join(notes).upper()

    sig1_path = select_signature(request.sig1, settings.sig1_dir)
    sig2_path = select_signature(request.sig2, settings.sig2_dir)
    add_signature(ws, sig1_path, "C42")
    add_signature(ws, sig2_path, "H42")

    wb.save(excel_path)
    pdf_path: Optional[str] = None
    if request.include_pdf:
        pdf_result = convert_excel_to_pdf(excel_path, folder)
        if pdf_result:
            pdf_path = str(pdf_result)
        else:
            message = "Loadsheet generated (Excel only - PDF conversion failed or disabled)"
    else:
        message = "Loadsheet generated (Excel only - PDF conversion skipped per request)"

    if request.include_pdf and pdf_path:
        message = "Loadsheet generated successfully"

    return GenerateResponse(
        excel_path=str(excel_path),
        pdf_path=str(pdf_path) if pdf_path else None,
        message=message,
        week_folder=format_folder_date(week_end),
    )
