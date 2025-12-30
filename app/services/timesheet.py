"""
Timesheet generation service.

Handles Excel template population and PDF conversion for weekly timesheets.
"""

import logging
from datetime import datetime, timedelta
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.cell import MergedCell

from ..config import get_settings
from ..schemas import DayModel, LoadModel, TimesheetRequest, GenerateResponse
from .helpers import (
    convert_excel_to_pdf,
    ensure_output_folder,
    format_date_for_cell,
    format_folder_date,
    get_week_end_from_date,
    sanitize_filename,
)

logger = logging.getLogger("paperworkgen.services.timesheet")
settings = get_settings()

DAY_ORDER = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]

DAY_MAPPINGS = {
    "monday": {"base_row": 8, "time_row": 8},
    "tuesday": {"base_row": 11, "time_row": 11},
    "wednesday": {"base_row": 14, "time_row": 14},
    "thursday": {"base_row": 17, "time_row": 17},
    "friday": {"base_row": 20, "time_row": 20},
    "saturday": {"base_row": 23, "time_row": 23},
    "sunday": {"base_row": 26, "time_row": 26},
}


def _calculate_total_hours(days: List[DayModel]) -> float:
    total = 0.0
    for day in days:
        if not day.total_hours:
            continue
        value = day.total_hours.strip()
        try:
            total += float(value)
        except ValueError:
            continue
    return total


def _get_day_date(week_end: datetime, day_name: str) -> datetime:
    index = DAY_ORDER.index(day_name) if day_name in DAY_ORDER else 0
    return week_end - timedelta(days=6 - index)


def _format_load(load: LoadModel) -> dict:
    return {
        "customer": load.customer.upper(),
        "car_count": str(load.car_count),
        "collection": load.collection.upper(),
        "delivery": load.delivery.upper(),
        "note": load.note.upper() if load.note else "",
    }


def _write_load_row(ws, row: int, load_data: dict) -> None:
    ws[f"C{row}"] = load_data["customer"]
    ws[f"D{row}"] = load_data["car_count"]
    ws[f"E{row}"] = load_data["collection"]
    ws[f"F{row}"] = load_data["delivery"]
    if load_data["note"]:
        ws[f"G{row}"] = load_data["note"]


def generate_timesheet(request: TimesheetRequest) -> GenerateResponse:
    """Generate timesheet Excel and PDF files."""

    if not settings.timesheet_template.exists():
        raise FileNotFoundError(f"Timesheet template not found: {settings.timesheet_template}")

    week_end_input = datetime.strptime(request.week_ending, "%Y-%m-%d")
    week_end = get_week_end_from_date(week_end_input)
    folder = ensure_output_folder(week_end)

    filename = f"timesheet_{week_end.strftime('%Y-%m-%d')}_{sanitize_filename(request.driver)}"
    excel_path = folder / f"{filename}.xlsx"

    wb = load_workbook(settings.timesheet_template)
    ws = wb.active

    ws["E5"] = format_date_for_cell(week_end)
    driver_cell = ws["K3"]
    if isinstance(driver_cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if driver_cell.coordinate in merged_range:
                top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left.value = request.driver.upper()
                break
        else:
            ws.cell(row=3, column=11).value = request.driver.upper()
    else:
        driver_cell.value = request.driver.upper()
    fleet_regs = ", ".join(dict.fromkeys([reg.upper() for reg in request.fleet_reg]))
    ws["K5"] = fleet_regs
    ws["H4"] = request.start_mileage
    ws["H5"] = request.end_mileage

    overflow_row = 29
    for day_data in request.days:
        day_name = day_data.day.lower()
        mapping = DAY_MAPPINGS.get(day_name)
        if not mapping:
            continue

        time_row = mapping["time_row"]
        start_time = (day_data.start_time or "").strip()
        finish_time = (day_data.finish_time or "").strip()
        total_hours = (day_data.total_hours or "").strip()

        if start_time.upper() in {"SICK", "HOLIDAY"} or finish_time.upper() in {"SICK", "HOLIDAY"}:
            marker = start_time.upper() if start_time else finish_time.upper()
            start_time = marker
            finish_time = marker
            total_hours = "0"

        ws[f"H{time_row}"] = start_time
        ws[f"I{time_row}"] = finish_time
        ws[f"J{time_row}"] = total_hours

        loads = day_data.loads or []
        for load_idx, load in enumerate(loads[:3]):
            row = mapping["base_row"] + load_idx
            if isinstance(load, dict) and load.get("message"):
                ws[f"C{row}"] = load["message"].upper()
                continue

            if isinstance(load, LoadModel):
                _write_load_row(ws, row, _format_load(load))
            elif isinstance(load, dict):
                if "customer" in load or "collection" in load:
                    load_data = {
                        "customer": str(load.get("customer", "")).upper(),
                        "car_count": str(load.get("car_count", "")),
                        "collection": str(load.get("collection", "")).upper(),
                        "delivery": str(load.get("delivery", "")).upper(),
                        "note": str(load.get("note", "") or load.get("custom_note", "")).upper(),
                    }
                    _write_load_row(ws, row, load_data)

        if len(loads) > 3:
            day_date = _get_day_date(week_end, day_name)
            for load in loads[3:]:
                if isinstance(load, dict) and load.get("message"):
                    continue
                if isinstance(load, LoadModel):
                    load_data = _format_load(load)
                elif isinstance(load, dict):
                    load_data = {
                        "customer": str(load.get("customer", "")).upper(),
                        "car_count": str(load.get("car_count", "")),
                        "collection": str(load.get("collection", "")).upper(),
                        "delivery": str(load.get("delivery", "")).upper(),
                        "note": str(load.get("note", "") or load.get("custom_note", "")).upper(),
                    }
                else:
                    continue

                ws[f"B{overflow_row}"] = day_date.strftime("%d/%m/%y")
                _write_load_row(ws, overflow_row, load_data)
                overflow_row += 1

    total_hours_value = request.weekly_total_hours.strip() if request.weekly_total_hours else ""
    if total_hours_value:
        ws["J29"] = total_hours_value
    else:
        weekly_total = _calculate_total_hours(request.days)
        ws["J29"] = str(weekly_total)

    wb.save(excel_path)

    pdf_path: Optional[str] = None
    if request.include_pdf:
        pdf_result = convert_excel_to_pdf(excel_path, folder)
        if pdf_result:
            pdf_path = str(pdf_result)
        else:
            message = "Timesheet generated (Excel only - PDF conversion failed or disabled)"
    else:
        message = "Timesheet generated (Excel only - PDF conversion skipped per request)"

    if request.include_pdf and pdf_path:
        message = "Timesheet generated successfully"

    return GenerateResponse(
        excel_path=str(excel_path),
        pdf_path=str(pdf_path) if pdf_path else None,
        message=message,
        week_folder=format_folder_date(week_end),
    )
